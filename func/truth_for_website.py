import pandas as pd
import math
from .data.truth import TruthFinderModel
from openpyxl import load_workbook, Workbook


def change_excel(file_origin, file_out):
    """
    将数据改变成置信度计算的格式并删除空项
    :param file_origin:
    :param file_out:
    :return:
    """
    book_origin = load_workbook(file_origin)
    sheet_origin = book_origin.active
    book = Workbook()
    sheet = book.active
    row_len = sheet_origin.max_row
    count = 2
    sheet.append(['website', 'fact', 'object'])
    for r in range(2, row_len + 1):
        value = sheet_origin.cell(r, 3).value
        if not (value == "" or value is None):
            if type(value) is int:
                value = str(value)
            sheet.cell(count, 1).value = sheet_origin.cell(r, 1).value
            sheet.cell(count, 2).value = value
            sheet.cell(count, 3).value = sheet_origin.cell(r, 2).value
            count += 1
    book.save(file_out)


def calculate_truth(file_origin, file_out):
    """
    计算置信度
    :param file_origin:
    :param file_out:
    :return:
    """
    df = pd.read_excel(file_origin)
    tf = TruthFinderModel()
    df = tf.train(df)
    df.to_excel(file_out)


def return_5truth(file_origin):
    """
    返回五个测绘引擎的置信度
    如果已有记录，直接查询返回
    如果没有，则进行计算
    :param file_origin: calculate_truth的输出文件
    :return:
    """
    try:
        with open("func/data/truth_5.txt", "r", encoding="utf-8") as file:
            # print(file.readline())
            value = eval(file.readline())
            return value
    except Exception as e:
        book = load_workbook(file_origin)
        sheet = book.active
        row_len = sheet.max_row
        result = {}
        for r in range(2, row_len+1):
            website = sheet.cell(r, 2).value
            if website not in result:
                value = sheet.cell(r, 6).value
                result[website] = 1-math.exp(-value)
        with open("data/truth_5.txt", "w", encoding="utf-8") as file:
            file.write(str(result))
        return result


def truth_for_website(flag=False):
    """
    注意flag若为True，耗时可能较长，并确保是否存在训练数据
    :param flag: 是否强制重新计算置信度
    :return:
    """
    if flag:
        change_excel("data/2000_all.xlsx", "../data/change.xlsx")
        calculate_truth("data/change.xlsx", "../data/out.xlsx")
    return return_5truth("data/out.xlsx")


if __name__ == "__main__":
    value = truth_for_website()
    print(value)
